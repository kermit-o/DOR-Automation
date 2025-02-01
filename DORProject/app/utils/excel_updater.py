from openpyxl import load_workbook
from datetime import datetime
from .xml_parser import extract_data_from_xml

def update_dor_excel(dor_file, xml_files):
    """Actualiza el archivo Excel DOR con los datos de los archivos XML."""
    wb = load_workbook(dor_file)
    ws = wb.active
    
    # Actualizar la fecha en D2
    ws["D2"] = datetime.today().strftime("%d-%m-%Y")

    # Mapeo de archivos XML a columnas
    column_map = {
        0: "C",  # archivo1
        1: "D",  # archivo2
        2: "E",  # archivo3
        3: "F"   # archivo4
    }

    # Mapeo de valores en celdas iniciales
    cell_map = {
        "REVENUE": "6",
        "NO_ROOMS": "8",
        "COMPLIMENTARY_ROOMS": "10",
        "HOUSE_USE_ROOMS": "11",
        "SUMOOO_ROOMSPERREPORT": "12"
    }

    # Mapeo de posiciones para reemplazo de datos antiguos
    replacement_map = {
        "6": "16",
        "8": "18",
        "10": "20",
        "11": "21",
        "12": "22"
    }

    # Procesar cada archivo XML y actualizar los valores en el Excel
    for i, xml_file in enumerate(xml_files):
        if i >= len(column_map):  
            break  

        column = column_map[i]
        data = extract_data_from_xml(xml_file)

        for key, row in cell_map.items():
            new_row = replacement_map[row]  # Obtener la nueva posición
            old_cell = f"{column}{row}"     # Celda actual
            new_cell = f"{column}{new_row}" # Nueva celda

            # Mover el dato existente a la nueva posición
            ws[new_cell] = ws[old_cell].value  
            # Insertar el nuevo valor en la celda original
            ws[old_cell] = data[key]  

    # Guardar el archivo Excel con los cambios
    wb.save(dor_file)
    return "✅ Archivo DOR actualizado correctamente."
