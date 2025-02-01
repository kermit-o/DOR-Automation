import openpyxl
from datetime import datetime
from flask_login import login_required

@login_required
def update_dor_template(dor_path, extracted_data_list):
    
    """Modifica la plantilla DOR con los datos extraídos"""
    wb = openpyxl.load_workbook(dor_path)
    ws = wb.active

    # 🔹 Modificar la fecha en D2
    ws["D2"] = datetime.today().strftime("%d-%m-%Y")

    # 🔹 Mapeo de celdas a reemplazar
    cell_mapping = {
        "C6": "C16", "C8": "C18", "C10": "C20", "C11": "C21", "C12": "C22",
        "D6": "D16", "D8": "D18", "D10": "D20", "D11": "D21", "D12": "D22",
        "E6": "E16", "E8": "E18", "E10": "E20", "E11": "E21", "E12": "E22",
        "F6": "F16", "F8": "F18", "F10": "F20", "F11": "F21", "F12": "F22"
    }

    for source, destination in cell_mapping.items():
        ws[destination] = ws[source].value  # Copiar valores

    # 🔹 Ingresar datos desde los XML
    columns = ["C", "D", "E", "F"]  # Corresponden a los archivos 1, 2, 3, 4
    keys = ["REVENUE", "NO_ROOMS", "COMPLIMENTARY_ROOMS", "HOUSE_USE_ROOMS", "SUMOOO_ROOMSPERREPORT"]
    rows = [6, 8, 10, 11, 12]

    for i, data in enumerate(extracted_data_list):
        col = columns[i]  # Determinar la columna a usar
        for key, row in zip(keys, rows):
            ws[f"{col}{row}"] = data[key]  # Asignar valores

    # 🔹 Guardar el archivo actualizado
    updated_path = dor_path.replace(".xlsx", "_updated.xlsx")
    wb.save(updated_path)
    return updated_path
