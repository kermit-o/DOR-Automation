from flask import Blueprint, render_template, request, redirect, url_for, send_from_directory, jsonify, flash
from openpyxl import load_workbook
import pandas as pd
import os
import json
from datetime import datetime
from .models_backup import DORReport, TemplateFile  # Import DORReport and TemplateFile from models
from app import db
import chardet  # Import para detectar codificación
from reportlab.pdfgen import canvas
import PyPDF2

UPLOAD_FOLDER = 'uploads'
os.makedirs(UPLOAD_FOLDER, exist_ok=True)

# Define el Blueprint antes de usarlo
main = Blueprint('main', __name__, template_folder='../templates')

@main.route('/')
def index():
    return render_template('upload.html')

@main.route('/reports', methods=['GET', 'POST'])
def list_reports():
    if request.method == 'POST':
        filter_filename = request.form.get('filename')
        if filter_filename:
            reports = DORReport.query.filter(DORReport.filename.contains(filter_filename)).all()
        else:
            reports = DORReport.query.all()
    else:
        reports = DORReport.query.all()

    return render_template('reports.html', reports=reports)

@main.route('/download/<filename>')
def download_file(filename):
    return send_from_directory(UPLOAD_FOLDER, filename, as_attachment=True)

@main.route('/files')
def list_files():
    # List all files in the upload folder
    files = os.listdir(UPLOAD_FOLDER)
    return render_template('files.html', files=files)

@main.route('/delete/<filename>', methods=['POST'])
def delete_file(filename):
    file_path = os.path.join(UPLOAD_FOLDER, filename)
    try:
        os.remove(file_path)
        flash(f"File {filename} deleted successfully.", "success")
    except Exception as e:
        flash(f"Error deleting file {filename}: {str(e)}", "error")
    return redirect(url_for('main.list_files'))

def detect_encoding(file_path):
    with open(file_path, 'rb') as f:
        result = chardet.detect(f.read(10000))  # Leer los primeros 10KB del archivo
    return result['encoding']

def extract_data_from_pdf(file_path):
    data = []
    try:
        with open(file_path, 'rb') as file:
            reader = PyPDF2.PdfReader(file)
            if not reader.pages:
                print(f"No pages found in PDF: {file_path}")
                return None
            for page in reader.pages:
                text = page.extract_text()
                if text:
                    data.append({"filename": os.path.basename(file_path), "text": text})
                else:
                    print(f"No text found on page {reader.pages.index(page) + 1} of PDF: {file_path}")
        return data if data else None
    except Exception as e:
        print(f"Error reading PDF file {file_path}: {str(e)}")
        return None

@main.route('/upload', methods=['POST'])
def upload_dor_file():
    if 'files' not in request.files:
        return "No file part", 400

    files = request.files.getlist('files')
    if not files:
        return "No selected files", 400

    is_template = 'is_template' in request.form

    combined_data = []
    for file in files:
        filename = file.filename
        if '.' not in filename or filename.rsplit('.', 1)[1].lower() not in ['xlsx', 'xlsm', 'csv', 'pdf']:
            return f"Unsupported file format: {filename}. Please upload a .xlsx, .xlsm, .csv, or .pdf file.", 400

        # Generar un nombre único para el archivo
        unique_filename = generate_unique_filename(UPLOAD_FOLDER, filename)
        file_path = os.path.join(UPLOAD_FOLDER, unique_filename)
        file.save(file_path)

        if is_template:
            # Guardar archivo de plantilla en la base de datos
            template_file = TemplateFile(filename=unique_filename)
            db.session.add(template_file)
            db.session.commit()
            return f"Template file uploaded successfully and saved as {unique_filename}.", 200

        try:
            data = None
            if unique_filename.endswith(('.xlsx', '.xlsm')):
                df = pd.read_excel(file_path)
                data = df.to_dict(orient='records')
            elif unique_filename.endswith('.csv'):
                encoding = detect_encoding(file_path)
                with open(file_path, 'r', encoding=encoding, errors='ignore') as f:
                    df = pd.read_csv(f, on_bad_lines='skip')
                data = df.to_dict(orient='records')
            elif unique_filename.endswith('.pdf'):
                data = extract_data_from_pdf(file_path)

            if data is not None:
                combined_data.extend(data)
            else:
                return f"Error processing file {unique_filename}: No text extracted from PDF.", 500

        except Exception as e:
            print(f"Error processing file {unique_filename}: {str(e)}")
            return f"Error processing file {unique_filename}: {str(e)}", 500

    # Obtener la última plantilla subida de la base de datos
    template_file = TemplateFile.query.order_by(TemplateFile.upload_date.desc()).first()
    if template_file:
        template_path = os.path.join(UPLOAD_FOLDER, template_file.filename)
        updated_file_path = update_template(combined_data, template_path)
        return f"Files uploaded and processed successfully! <a href='/download/{os.path.basename(updated_file_path)}'>Download Updated Template</a>"
    else:
        return "Files uploaded successfully, but no template file found to process data.", 200

def generate_unique_filename(directory, filename):
    base, extension = os.path.splitext(filename)
    timestamp = datetime.now().strftime("%Y%m%d%H%M%S")
    unique_filename = f"{base}_{timestamp}{extension}"
    return unique_filename

def update_template(data, template_path):
    # Imprimir la ruta absoluta del archivo de plantilla
    print(f"Ruta absoluta del archivo de plantilla: {template_path}")
    
    # Load the template file
    workbook = load_workbook(template_path, keep_vba=True)

    # Verificar y listar todas las hojas disponibles
    print(f"Hojas disponibles en el archivo de plantilla: {workbook.sheetnames}")

    # Assuming we need to update a specific sheet in the template
    sheet_name = 'Set-up'  # Cambiar al nombre correcto de la hoja deseada
    if sheet_name not in workbook.sheetnames:
        return f"Error: La hoja '{sheet_name}' no existe en el archivo de plantilla.", 500

    sheet = workbook[sheet_name]

    # Clear existing data in the sheet (optional)
    merged_cells = list(sheet.merged_cells.ranges)  # Copiar la lista de celdas combinadas
    for merged_cell in merged_cells:
        sheet.unmerge_cells(str(merged_cell))

    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
        for cell in row:
            cell.value = None

    # Write new data to the template
    for index, row in enumerate(data):
        for col_index, (key, value) in enumerate(row.items()):
            sheet.cell(row=index + 2, column=col_index + 1, value=value)

    # Save the updated template
    updated_file_path = os.path.join(UPLOAD_FOLDER, 'updated_template.xlsm')
    workbook.save(updated_file_path)
    return updated_file_path

def update_database(data, filename):
    for item in data:
        try:
            item['filename'] = filename
            print(f"Processing item: {item}")

            existing_record = db.session.query(DORReport).filter_by(filename=item['filename']).first()
            if existing_record:
                existing_record.processed_data = item.get('text', existing_record.processed_data)
                existing_record.upload_date = datetime.utcnow()
                db.session.commit()
            else:
                new_report = DORReport(
                    filename=item['filename'],
                    upload_date=datetime.utcnow(),
                    processed_data=item.get('text', '')
                )
                db.session.add(new_report)
                db.session.commit()
        except Exception as e:
            print(f"Error updating database record: {str(e)}")

@main.route('/export/<int:report_id>')
def export_report(report_id):
    report = DORReport.query.get_or_404(report_id)

    pdf_path = os.path.join(UPLOAD_FOLDER, f"report_{report.id}.pdf")
    c = canvas.Canvas(pdf_path)
    c.drawString(100, 800, f"Report ID: {report.id}")
    c.drawString(100, 780, f"Filename: {report.filename}")
    c.drawString(100, 760, f"Upload Date: {report.upload_date}")
    c.drawString(100, 740, "Processed Data:")

    y_position = 720
    data = json.loads(report.processed_data)
    for row in data:
        c.drawString(100, y_position, str(row))
        y_position -= 20
        if y_position < 50:
            c.showPage()
            y_position = 800

    c.save()

    return f"PDF generated successfully! <a href='/download/{os.path.basename(pdf_path)}'>Download PDF</a>"